# =============================================================================
# xlsx_engine.py — Advanced Excel Generation Engine v1.0
# =============================================================================
# 3-layer architecture (mirrors pptx_engine.py):
#   Layer 1: AI Planner (Opus) — plans workbook structure from content
#   Layer 2: Validator — enforces quality rules deterministically
#   Layer 3: Renderer — builds .xlsx with openpyxl
#
# Features: multi-sheet, native data types, formulas (SUM/AVG/COUNT),
#           charts, conditional formatting, auto-filter, freeze panes,
#           hyperlinks, Millennium BCP branding.
# =============================================================================

import io
import json
import logging
import re
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from config import EXPORT_BRAND_COLOR, EXPORT_AGENT_NAME, EXPORT_BRAND_NAME, APP_VERSION

logger = logging.getLogger(__name__)

# =============================================================================
# BRAND CONSTANTS
# =============================================================================
BRAND_ACCENT_HEX = str(EXPORT_BRAND_COLOR or "#DE3163").strip().lstrip("#").upper()
if len(BRAND_ACCENT_HEX) != 6:
    BRAND_ACCENT_HEX = "DE3163"
BRAND_ACCENT_LIGHTER = "FCE4EC"  # Light pink for subtle highlights
BRAND_FONT = "Montserrat"
FALLBACK_FONT = "Calibri"
ZEBRA_HEX = "F5F5F5"
HEADER_FONT_COLOR = "FFFFFF"

# =============================================================================
# VALIDATION CONSTANTS
# =============================================================================
MAX_SHEETS = 20
MAX_ROWS_PER_SHEET = 10000
MAX_COLS = 50
MAX_SHEET_NAME_LEN = 31
MAX_CHART_SERIES = 8
MAX_FORMULA_COLS = 20

# =============================================================================
# DATA TYPE DETECTION
# =============================================================================

_DATE_PATTERNS = [
    re.compile(r"^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}"),     # ISO datetime
    re.compile(r"^\d{4}-\d{2}-\d{2}$"),                     # ISO date
    re.compile(r"^\d{2}/\d{2}/\d{4}$"),                     # DD/MM/YYYY
    re.compile(r"^\d{2}-\d{2}-\d{4}$"),                     # DD-MM-YYYY
]

_PERCENT_PATTERN = re.compile(r"^[+-]?\d+(?:[.,]\d+)?%$")
_NUMBER_PATTERN = re.compile(r"^[+-]?\d+(?:[.,]\d+)?$")
_CURRENCY_PATTERN = re.compile(r"^[€$£]\s*\d+(?:[.,]\d+)?$|^\d+(?:[.,]\d+)?\s*[€$£]$")
_URL_PATTERN = re.compile(r"^https?://\S+$", re.IGNORECASE)


def detect_column_type(values: List[str], col_name: str = "") -> str:
    """Detect the predominant data type of a column from sample values.

    Returns one of: 'number', 'date', 'percent', 'currency', 'url', 'text'.
    """
    if not values:
        return "text"

    # Sample up to 50 non-empty values
    samples = [str(v).strip() for v in values if str(v).strip()][:50]
    if not samples:
        return "text"

    # Column name hints
    name_lower = col_name.lower().replace("_", " ")
    if any(kw in name_lower for kw in ("url", "link", "href")):
        return "url"
    if any(kw in name_lower for kw in ("date", "data", "created", "updated", "timestamp")):
        return "date"
    if any(kw in name_lower for kw in ("percent", "percentagem", "taxa", "%")):
        return "percent"

    type_counts = {"number": 0, "date": 0, "percent": 0, "currency": 0, "url": 0, "text": 0}

    for val in samples:
        if _PERCENT_PATTERN.match(val):
            type_counts["percent"] += 1
        elif _CURRENCY_PATTERN.match(val):
            type_counts["currency"] += 1
        elif _URL_PATTERN.match(val):
            type_counts["url"] += 1
        elif any(p.match(val) for p in _DATE_PATTERNS):
            type_counts["date"] += 1
        elif _NUMBER_PATTERN.match(val):
            type_counts["number"] += 1
        else:
            type_counts["text"] += 1

    # Need >60% consensus for non-text type
    threshold = len(samples) * 0.6
    for dtype in ("number", "date", "percent", "currency", "url"):
        if type_counts[dtype] >= threshold:
            return dtype

    return "text"


def _parse_number(value: str) -> Optional[float]:
    """Parse a string to float, handling comma decimals."""
    txt = str(value or "").strip()
    if not txt:
        return None
    # Remove currency symbols
    txt = re.sub(r"[€$£\s]", "", txt)
    # Remove % for percent
    txt = txt.rstrip("%")
    # Handle comma as decimal separator (European)
    if "," in txt and "." not in txt:
        txt = txt.replace(",", ".")
    elif "," in txt and "." in txt:
        # 1.234,56 format → remove dots, comma→dot
        txt = txt.replace(".", "").replace(",", ".")
    try:
        return float(txt)
    except (ValueError, TypeError):
        return None


def _parse_date(value: str) -> Optional[datetime]:
    """Parse a string to datetime."""
    txt = str(value or "").strip()
    if not txt:
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%d %H:%M", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(txt[:19], fmt)
        except (ValueError, TypeError):
            continue
    return None


# =============================================================================
# LAYER 3: RENDERER
# =============================================================================

def _safe_sheet_title(title: str, existing: List[str] = None) -> str:
    """Sanitize and deduplicate sheet title."""
    if not title:
        title = "Sheet"
    forbidden = set("[]:*?/\\")
    safe = "".join("_" if ch in forbidden else ch for ch in str(title))
    safe = safe.strip().strip("'")
    if not safe:
        safe = "Sheet"
    safe = safe[:MAX_SHEET_NAME_LEN]

    if existing:
        base = safe
        counter = 1
        while safe in existing:
            suffix = f" ({counter})"
            safe = base[:MAX_SHEET_NAME_LEN - len(suffix)] + suffix
            counter += 1

    return safe


def _apply_branding(ws, headers: List[str], *, title: str = "", row_count: int = 0):
    """Apply Millennium BCP branding to a worksheet header area."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    brand_fill = PatternFill(start_color=BRAND_ACCENT_HEX, end_color=BRAND_ACCENT_HEX, fill_type="solid")
    header_font = Font(name=FALLBACK_FONT, bold=True, color=HEADER_FONT_COLOR, size=11)
    border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    ncols = len(headers) if headers else 1

    # Row 1: Title
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(ncols, 1))
        title_cell = ws.cell(row=1, column=1, value=f"{EXPORT_AGENT_NAME} — {title}")
        title_cell.font = Font(name=FALLBACK_FONT, bold=True, size=14, color=BRAND_ACCENT_HEX)

    # Row 2: Subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(ncols, 1))
    sub = ws.cell(
        row=2, column=1,
        value=f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} | {row_count} registos",
    )
    sub.font = Font(name=FALLBACK_FONT, size=9, color="666666", italic=True)

    # Row 3: spacer
    # Row 4: Headers
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=_clean_header(h))
        cell.font = header_font
        cell.fill = brand_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    return 5  # data starts at row 5


def _clean_header(h: str) -> str:
    """Clean header for display."""
    return h.replace("_", " ").title()


def _write_data_rows(
    ws,
    rows: List[Dict[str, Any]],
    columns: List[str],
    col_types: Dict[str, str],
    start_row: int = 5,
) -> int:
    """Write data rows with native types and styling. Returns next empty row."""
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers

    data_font = Font(name=FALLBACK_FONT, size=10)
    zebra_fill = PatternFill(start_color=ZEBRA_HEX, end_color=ZEBRA_HEX, fill_type="solid")
    border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    for r_idx, row in enumerate(rows):
        excel_row = start_row + r_idx
        is_zebra = r_idx % 2 == 1

        for c_idx, col in enumerate(columns, 1):
            raw_val = row.get(col, "")
            cell = ws.cell(row=excel_row, column=c_idx)
            cell.font = data_font
            cell.border = border
            if is_zebra:
                cell.fill = zebra_fill

            dtype = col_types.get(col, "text")

            if dtype == "number":
                num = _parse_number(str(raw_val))
                if num is not None:
                    cell.value = num
                    # Format: integer vs decimal
                    if num == int(num) and abs(num) < 1e12:
                        cell.number_format = "#,##0"
                    else:
                        cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.value = str(raw_val)

            elif dtype == "percent":
                num = _parse_number(str(raw_val))
                if num is not None:
                    # If value is already 0-100 range, divide by 100
                    if abs(num) > 1:
                        num = num / 100
                    cell.value = num
                    cell.number_format = "0.0%"
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.value = str(raw_val)

            elif dtype == "currency":
                num = _parse_number(str(raw_val))
                if num is not None:
                    cell.value = num
                    cell.number_format = '#,##0.00 €'
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.value = str(raw_val)

            elif dtype == "date":
                dt = _parse_date(str(raw_val))
                if dt is not None:
                    cell.value = dt
                    cell.number_format = "DD/MM/YYYY"
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.value = str(raw_val)

            elif dtype == "url":
                url_str = str(raw_val).strip()
                if url_str and _URL_PATTERN.match(url_str):
                    cell.value = url_str
                    cell.hyperlink = url_str
                    cell.font = Font(name=FALLBACK_FONT, size=10, color="0563C1", underline="single")
                else:
                    cell.value = str(raw_val)

            else:
                cell.value = str(raw_val)

    return start_row + len(rows)


def _add_formulas(ws, columns: List[str], col_types: Dict[str, str],
                  data_start_row: int, data_end_row: int):
    """Add SUM/AVG/COUNT formulas row below data for numeric columns."""
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    if data_end_row <= data_start_row:
        return data_end_row

    formula_row = data_end_row + 1
    summary_fill = PatternFill(start_color=BRAND_ACCENT_LIGHTER, end_color=BRAND_ACCENT_LIGHTER, fill_type="solid")
    bold_font = Font(name=FALLBACK_FONT, bold=True, size=10)
    border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="medium", color=BRAND_ACCENT_HEX),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    has_formulas = False
    for c_idx, col in enumerate(columns, 1):
        dtype = col_types.get(col, "text")
        col_letter = get_column_letter(c_idx)
        cell = ws.cell(row=formula_row, column=c_idx)
        cell.fill = summary_fill
        cell.font = bold_font
        cell.border = border

        if dtype in ("number", "currency"):
            cell.value = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row - 1})"
            if dtype == "currency":
                cell.number_format = '#,##0.00 €'
            else:
                cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right")
            has_formulas = True
        elif dtype == "percent":
            cell.value = f"=AVERAGE({col_letter}{data_start_row}:{col_letter}{data_end_row - 1})"
            cell.number_format = "0.0%"
            cell.alignment = Alignment(horizontal="right")
            has_formulas = True
        elif c_idx == 1:
            cell.value = "TOTAL"

    # Row for count
    if has_formulas:
        count_row = formula_row + 1
        for c_idx, col in enumerate(columns, 1):
            dtype = col_types.get(col, "text")
            col_letter = get_column_letter(c_idx)
            cell = ws.cell(row=count_row, column=c_idx)
            cell.font = Font(name=FALLBACK_FONT, italic=True, size=9, color="666666")

            if dtype in ("number", "currency", "percent"):
                cell.value = f"=COUNTA({col_letter}{data_start_row}:{col_letter}{data_end_row - 1})"
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            elif c_idx == 1:
                cell.value = "Contagem"

    return formula_row + (2 if has_formulas else 1)


def _auto_column_widths(ws, columns: List[str], col_types: Dict[str, str],
                        rows: List[Dict[str, Any]]):
    """Set column widths based on content sampling."""
    from openpyxl.utils import get_column_letter

    for c_idx, col in enumerate(columns, 1):
        header_len = len(_clean_header(col))
        max_len = header_len

        # Sample up to 50 rows
        for row in rows[:50]:
            val = str(row.get(col, ""))
            max_len = max(max_len, len(val))

        dtype = col_types.get(col, "text")
        if dtype == "url":
            width = min(max_len + 4, 40)
        elif dtype in ("number", "currency", "percent"):
            width = min(max(max_len + 4, 12), 20)
        elif dtype == "date":
            width = max(12, header_len + 4)
        else:
            width = min(max_len + 4, 50)

        ws.column_dimensions[get_column_letter(c_idx)].width = max(width, 8)


def _add_auto_filter_and_freeze(ws, columns: List[str], header_row: int = 4):
    """Add auto-filter on header row and freeze panes below it."""
    from openpyxl.utils import get_column_letter

    if not columns:
        return

    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row}"
    ws.freeze_panes = f"A{header_row + 1}"


def _add_chart_sheet(
    wb,
    data_ws,
    chart_spec: Dict[str, Any],
    columns: List[str],
    col_types: Dict[str, str],
    data_start_row: int,
    data_end_row: int,
    existing_names: List[str],
):
    """Add a chart to the workbook based on chart_spec."""
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.utils import get_column_letter

    chart_type = chart_spec.get("type", "bar").lower()
    chart_title = chart_spec.get("title", "Gráfico")
    label_col = chart_spec.get("label_column", "")
    value_cols = chart_spec.get("value_columns", [])

    if not value_cols or not label_col:
        # Auto-detect: first text col as label, first numeric cols as values
        for col in columns:
            if col_types.get(col) == "text" and not label_col:
                label_col = col
            elif col_types.get(col) in ("number", "currency", "percent") and len(value_cols) < MAX_CHART_SERIES:
                value_cols.append(col)

    if not label_col or not value_cols:
        return None

    label_idx = columns.index(label_col) + 1 if label_col in columns else None
    value_indices = [(columns.index(vc) + 1) for vc in value_cols if vc in columns]

    if not label_idx or not value_indices:
        return None

    nrows = data_end_row - data_start_row

    if chart_type == "pie":
        chart = PieChart()
        chart.style = 10
        if value_indices:
            data_ref = Reference(data_ws, min_col=value_indices[0], min_row=data_start_row - 1,
                                 max_row=data_start_row + nrows - 1)
            cats_ref = Reference(data_ws, min_col=label_idx, min_row=data_start_row,
                                 max_row=data_start_row + nrows - 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
    elif chart_type == "line":
        chart = LineChart()
        chart.style = 10
        cats_ref = Reference(data_ws, min_col=label_idx, min_row=data_start_row,
                             max_row=data_start_row + nrows - 1)
        for vi in value_indices:
            data_ref = Reference(data_ws, min_col=vi, min_row=data_start_row - 1,
                                 max_row=data_start_row + nrows - 1)
            chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
    else:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        cats_ref = Reference(data_ws, min_col=label_idx, min_row=data_start_row,
                             max_row=data_start_row + nrows - 1)
        for vi in value_indices:
            data_ref = Reference(data_ws, min_col=vi, min_row=data_start_row - 1,
                                 max_row=data_start_row + nrows - 1)
            chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

    chart.title = chart_title
    chart.width = 20
    chart.height = 12

    # Add chart to its own sheet or embed in data sheet
    embed_in = chart_spec.get("embed_in_data_sheet", False)
    if embed_in:
        ws = data_ws
        # Place below data
        ws.add_chart(chart, f"A{data_end_row + 4}")
        return None
    else:
        sheet_name = _safe_sheet_title(f"Gráfico - {chart_title}", existing_names)
        chart_ws = wb.create_sheet(title=sheet_name)
        chart_ws.add_chart(chart, "A2")
        # Add title cell
        chart_ws.cell(row=1, column=1, value=chart_title)
        from openpyxl.styles import Font
        chart_ws["A1"].font = Font(name=FALLBACK_FONT, bold=True, size=14, color=BRAND_ACCENT_HEX)
        return sheet_name


def _add_conditional_formatting(ws, columns: List[str], col_types: Dict[str, str],
                                data_start_row: int, data_end_row: int):
    """Add conditional formatting for numeric columns (color scale)."""
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

    for c_idx, col in enumerate(columns, 1):
        dtype = col_types.get(col, "text")
        if dtype in ("number", "currency", "percent") and data_end_row > data_start_row:
            col_letter = get_column_letter(c_idx)
            cell_range = f"{col_letter}{data_start_row}:{col_letter}{data_end_row - 1}"
            rule = ColorScaleRule(
                start_type="min", start_color="FFFFFF",
                mid_type="percentile", mid_value=50, mid_color=BRAND_ACCENT_LIGHTER,
                end_type="max", end_color=BRAND_ACCENT_HEX,
            )
            ws.conditional_formatting.add(cell_range, rule)


def build_data_sheet(
    wb,
    sheet_spec: Dict[str, Any],
    existing_names: List[str],
) -> Optional[str]:
    """Build a single data sheet from a sheet spec.

    sheet_spec keys:
        name (str): sheet name
        title (str): display title in header
        data (list[dict]): rows as dicts
        columns (list[str]): column order
        formulas (bool): add SUM/AVG row, default True
        auto_filter (bool): add auto-filter + freeze, default True
        conditional (bool): add conditional formatting for numerics, default True
        chart (dict|None): optional chart spec to embed or create
    """
    name = sheet_spec.get("name", "Sheet")
    title = sheet_spec.get("title", name)
    data = sheet_spec.get("data", [])
    columns = sheet_spec.get("columns", [])
    add_formulas = sheet_spec.get("formulas", True)
    add_filter = sheet_spec.get("auto_filter", True)
    add_conditional = sheet_spec.get("conditional", False)
    chart_spec = sheet_spec.get("chart")

    if not data:
        return None

    # Infer columns from first row if not provided
    if not columns and isinstance(data[0], dict):
        columns = list(data[0].keys())
    if not columns:
        return None

    # Normalize rows to dicts
    normalized = []
    for row in data:
        if isinstance(row, dict):
            normalized.append({c: row.get(c, "") for c in columns})
        elif isinstance(row, (list, tuple)):
            normalized.append({c: (row[i] if i < len(row) else "") for i, c in enumerate(columns)})
    data = normalized

    if not data:
        return None

    # Detect column types
    col_types = {}
    for col in columns:
        values = [str(row.get(col, "")) for row in data[:50]]
        col_types[col] = detect_column_type(values, col)

    # Create sheet
    safe_name = _safe_sheet_title(name, existing_names)
    ws = wb.create_sheet(title=safe_name)

    # Apply branding header
    data_start = _apply_branding(ws, columns, title=title, row_count=len(data))

    # Write data with native types
    data_end = _write_data_rows(ws, data, columns, col_types, start_row=data_start)

    # Auto-filter + freeze
    if add_filter:
        _add_auto_filter_and_freeze(ws, columns, header_row=4)

    # Formulas
    if add_formulas:
        _add_formulas(ws, columns, col_types, data_start, data_end)

    # Conditional formatting
    if add_conditional:
        _add_conditional_formatting(ws, columns, col_types, data_start, data_end)

    # Auto-width
    _auto_column_widths(ws, columns, col_types, data)

    # Chart
    if chart_spec and isinstance(chart_spec, dict):
        _add_chart_sheet(
            wb, ws, chart_spec, columns, col_types,
            data_start, data_end, existing_names + [safe_name],
        )

    return safe_name


def build_summary_sheet(
    wb,
    summary_spec: Dict[str, Any],
    existing_names: List[str],
) -> Optional[str]:
    """Build a summary/dashboard sheet with KPIs and overview stats.

    summary_spec keys:
        name (str): sheet name, default "Resumo"
        title (str): display title
        kpis (list[dict]): each {label, value, description?}
        sections (list[dict]): each {title, items: list[str]}
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    name = summary_spec.get("name", "Resumo")
    title = summary_spec.get("title", "Resumo Executivo")
    kpis = summary_spec.get("kpis", [])
    sections = summary_spec.get("sections", [])

    if not kpis and not sections:
        return None

    safe_name = _safe_sheet_title(name, existing_names)
    ws = wb.create_sheet(title=safe_name)

    brand_fill = PatternFill(start_color=BRAND_ACCENT_HEX, end_color=BRAND_ACCENT_HEX, fill_type="solid")
    light_fill = PatternFill(start_color=BRAND_ACCENT_LIGHTER, end_color=BRAND_ACCENT_LIGHTER, fill_type="solid")

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name=FALLBACK_FONT, bold=True, size=18, color=BRAND_ACCENT_HEX)

    # Subtitle
    ws.merge_cells("A2:F2")
    ws.cell(row=2, column=1, value=f"{EXPORT_AGENT_NAME} | {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(
        name=FALLBACK_FONT, size=9, color="666666", italic=True)

    current_row = 4

    # KPIs in cards (2 per row)
    if kpis:
        ws.cell(row=current_row, column=1, value="Indicadores Chave").font = Font(
            name=FALLBACK_FONT, bold=True, size=13, color=BRAND_ACCENT_HEX)
        current_row += 1

        for i in range(0, len(kpis), 2):
            for offset, kpi in enumerate(kpis[i:i + 2]):
                col_start = 1 + offset * 3
                # Value cell
                val_cell = ws.cell(row=current_row, column=col_start, value=kpi.get("value", ""))
                val_cell.font = Font(name=FALLBACK_FONT, bold=True, size=28, color=BRAND_ACCENT_HEX)
                val_cell.alignment = Alignment(horizontal="center")
                ws.merge_cells(start_row=current_row, start_column=col_start,
                               end_row=current_row, end_column=col_start + 1)

                # Label cell
                label_cell = ws.cell(row=current_row + 1, column=col_start, value=kpi.get("label", ""))
                label_cell.font = Font(name=FALLBACK_FONT, bold=True, size=11)
                label_cell.alignment = Alignment(horizontal="center")
                ws.merge_cells(start_row=current_row + 1, start_column=col_start,
                               end_row=current_row + 1, end_column=col_start + 1)

                # Description
                desc = kpi.get("description", "")
                if desc:
                    desc_cell = ws.cell(row=current_row + 2, column=col_start, value=desc)
                    desc_cell.font = Font(name=FALLBACK_FONT, size=9, color="666666", italic=True)
                    desc_cell.alignment = Alignment(horizontal="center")
                    ws.merge_cells(start_row=current_row + 2, start_column=col_start,
                                   end_row=current_row + 2, end_column=col_start + 1)

                # Background
                for r in range(current_row, current_row + 3):
                    for c in range(col_start, col_start + 2):
                        ws.cell(row=r, column=c).fill = light_fill

            current_row += 4

    # Sections
    if sections:
        current_row += 1
        for section in sections:
            ws.cell(row=current_row, column=1, value=section.get("title", "")).font = Font(
                name=FALLBACK_FONT, bold=True, size=12, color=BRAND_ACCENT_HEX)
            current_row += 1

            for item in section.get("items", []):
                ws.cell(row=current_row, column=1, value=f"  • {item}").font = Font(
                    name=FALLBACK_FONT, size=10)
                current_row += 1

            current_row += 1

    # Set column widths
    for col in range(1, 7):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Freeze panes below title area for scrolling
    ws.freeze_panes = "A5"

    return safe_name


# =============================================================================
# LAYER 2: VALIDATOR
# =============================================================================

def _validate_and_fix_workbook(spec: Dict[str, Any]) -> Dict[str, Any]:
    """Validate and fix a workbook spec. Enforces quality rules deterministically.

    Rules:
    1. Max 20 sheets
    2. Max 10,000 rows per sheet (split if needed)
    3. Max 50 columns (trim)
    4. Sheet names must be unique and valid
    5. Empty sheets are removed
    6. Column names are deduplicated
    7. At least one data sheet required
    """
    sheets = spec.get("sheets", [])
    summary = spec.get("summary")
    charts = spec.get("charts", [])

    # ── Rule 5: Remove empty sheets ──
    sheets = [s for s in sheets if s.get("data") and len(s["data"]) > 0]

    # ── Rule 7: Ensure at least one sheet ──
    if not sheets and not summary:
        return spec

    # ── Rule 1: Max sheets ──
    if len(sheets) > MAX_SHEETS:
        sheets = sheets[:MAX_SHEETS]

    validated_sheets = []
    existing_names = []

    for sheet in sheets:
        data = sheet.get("data", [])
        columns = sheet.get("columns", [])

        # ── Rule 3: Max columns ──
        if len(columns) > MAX_COLS:
            columns = columns[:MAX_COLS]
            sheet["columns"] = columns
            # Trim data to match
            if data and isinstance(data[0], dict):
                col_set = set(columns)
                sheet["data"] = [{k: v for k, v in row.items() if k in col_set} for row in data]

        # ── Rule 6: Deduplicate column names ──
        seen = {}
        deduped = []
        for col in columns:
            if col in seen:
                seen[col] += 1
                deduped.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 0
                deduped.append(col)
        sheet["columns"] = deduped

        # ── Rule 2: Split large sheets ──
        if len(data) > MAX_ROWS_PER_SHEET:
            chunk_idx = 0
            for start in range(0, len(data), MAX_ROWS_PER_SHEET):
                chunk_idx += 1
                chunk = data[start:start + MAX_ROWS_PER_SHEET]
                split_sheet = {**sheet, "data": chunk}
                base_name = sheet.get("name", "Sheet")
                if chunk_idx > 1:
                    split_sheet["name"] = f"{base_name} ({chunk_idx})"
                    split_sheet["title"] = f"{sheet.get('title', base_name)} (cont. {chunk_idx})"
                validated_sheets.append(split_sheet)
        else:
            validated_sheets.append(sheet)

    # ── Rule 4: Unique sheet names ──
    for sheet in validated_sheets:
        name = sheet.get("name", "Sheet")
        safe = _safe_sheet_title(name, existing_names)
        sheet["name"] = safe
        existing_names.append(safe)

    # ── Rule 1 again after splits ──
    validated_sheets = validated_sheets[:MAX_SHEETS]

    spec["sheets"] = validated_sheets
    if charts:
        spec["charts"] = charts[:MAX_CHART_SERIES]
    return spec


# =============================================================================
# LAYER 1: AI PLANNER (Opus)
# =============================================================================

_WORKBOOK_PLANNER_PROMPT = """You are an expert Excel workbook architect for Millennium BCP (Portuguese bank).
Given content/data, plan a professional multi-sheet workbook structure.

RULES:
1. Output ONLY valid JSON — no markdown, no explanation
2. Output format: {"sheets": [...], "summary": {...} or null, "charts": [...]}
3. Each sheet: {"name": "...", "title": "...", "columns": [...], "data": [...], "formulas": true/false, "auto_filter": true, "conditional": true/false, "chart": {...} or null}
4. Summary sheet (optional): {"name": "Resumo", "title": "...", "kpis": [{"value": "...", "label": "...", "description": "..."}], "sections": [{"title": "...", "items": ["..."]}]}
5. Charts: {"type": "bar"|"line"|"pie", "title": "...", "label_column": "...", "value_columns": [...], "embed_in_data_sheet": true/false}
6. Sheet names max 31 chars, no special chars []:*?/\\
7. When data has natural groupings (by team, date, category), split into separate sheets
8. Always include formulas (SUM/AVG) for numeric columns
9. Use conditional formatting for performance metrics
10. KPI summary sheet when there are clear metrics to highlight
11. Max 20 sheets, max 50 columns per sheet
12. Column names should be clean and descriptive (snake_case)
13. Keep data values as-is — don't transform or summarize data
14. IMPORTANT: Preserve ALL data rows, do not drop or aggregate unless explicitly requested

Example output:
{
  "sheets": [
    {
      "name": "Dados Principais",
      "title": "User Stories Ativas",
      "columns": ["id", "title", "state", "assigned_to", "effort"],
      "data": [{"id": "123", "title": "Login", "state": "Active", "assigned_to": "João", "effort": "5"}],
      "formulas": true,
      "auto_filter": true,
      "conditional": true,
      "chart": {"type": "bar", "title": "Effort por Assignee", "label_column": "assigned_to", "value_columns": ["effort"], "embed_in_data_sheet": true}
    }
  ],
  "summary": {
    "name": "Resumo",
    "title": "Dashboard Executivo",
    "kpis": [{"value": "42", "label": "Total USs", "description": "User stories ativas"}],
    "sections": [{"title": "Destaques", "items": ["Sprint velocity: 35 pts/sprint"]}]
  },
  "charts": []
}"""


async def plan_workbook_with_opus(
    content: str,
    title: str = "",
    context: str = "",
    *,
    tier: str = "pro",
) -> Optional[Dict[str, Any]]:
    """Use Claude Opus to plan optimal workbook structure from content.

    Args:
        content: Free-text content or JSON data to structure into sheets
        title: Workbook title hint
        context: Additional context (conversation history, user preferences)
        tier: LLM tier to use (default "pro" = Opus)

    Returns:
        Workbook spec dict or None on failure
    """
    from llm_provider import llm_simple

    parts = [_WORKBOOK_PLANNER_PROMPT]
    if title:
        parts.append(f"\nWorkbook title: {title}")
    if context:
        parts.append(f"\nContext: {context[:2000]}")
    parts.append(f"\nContent to structure into Excel sheets:\n{content[:8000]}")

    full_prompt = "\n".join(parts)

    try:
        raw = await llm_simple(full_prompt, tier=tier, max_tokens=6000)
    except Exception as e:
        logger.warning("[XlsxEngine] Opus planning failed: %s", e)
        return None

    if not raw:
        return None

    # Parse JSON from response
    try:
        # Strip markdown fences
        cleaned = raw.strip()
        if cleaned.startswith("```"):
            first_newline = cleaned.index("\n")
            last_fence = cleaned.rfind("```")
            if last_fence > first_newline:
                cleaned = cleaned[first_newline + 1:last_fence].strip()

        # Find JSON object
        brace_start = cleaned.find("{")
        brace_end = cleaned.rfind("}")
        if brace_start >= 0 and brace_end > brace_start:
            cleaned = cleaned[brace_start:brace_end + 1]

        spec = json.loads(cleaned)
        if not isinstance(spec, dict):
            return None

        return spec
    except (json.JSONDecodeError, ValueError) as e:
        logger.warning("[XlsxEngine] Failed to parse Opus response: %s", e)
        return None


def _fallback_workbook_from_data(
    data: List[Dict[str, Any]],
    title: str = "Export",
    columns: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """Create a simple single-sheet workbook spec from flat data (fallback)."""
    if not data:
        return {"sheets": [], "summary": None, "charts": []}

    if not columns:
        if isinstance(data[0], dict):
            columns = list(data[0].keys())
        else:
            columns = [f"col_{i+1}" for i in range(len(data[0]) if data else 0)]

    return {
        "sheets": [{
            "name": title[:31],
            "title": title,
            "columns": columns,
            "data": data,
            "formulas": True,
            "auto_filter": True,
            "conditional": False,
            "chart": None,
        }],
        "summary": None,
        "charts": [],
    }


# =============================================================================
# MAIN ENTRY POINT
# =============================================================================

def generate_workbook(spec: Dict[str, Any]) -> io.BytesIO:
    """Generate an Excel workbook from a validated spec.

    This is the main renderer (Layer 3). It:
    1. Validates the spec through the validation layer
    2. Builds each sheet (data sheets, summary, charts)
    3. Returns the workbook as BytesIO

    Args:
        spec: Workbook specification dict with "sheets", "summary", "charts"

    Returns:
        io.BytesIO containing the .xlsx file
    """
    from openpyxl import Workbook

    # Apply validation
    spec = _validate_and_fix_workbook(spec)

    wb = Workbook()
    # Remove default empty sheet
    default_sheet = wb.active
    existing_names = []

    # Build summary sheet first (if present)
    summary = spec.get("summary")
    if summary and isinstance(summary, dict):
        result = build_summary_sheet(wb, summary, existing_names)
        if result:
            existing_names.append(result)

    # Build data sheets
    sheets = spec.get("sheets", [])
    data_worksheets = {}  # name → (ws, columns, col_types, data_start, data_end)

    for sheet_spec in sheets:
        result = build_data_sheet(wb, sheet_spec, existing_names)
        if result:
            existing_names.append(result)

    # Build standalone chart sheets
    charts = spec.get("charts", [])
    for chart_spec in charts:
        if not isinstance(chart_spec, dict):
            continue
        # Find the source data sheet
        source_sheet = chart_spec.get("source_sheet", "")
        if source_sheet and source_sheet in [ws.title for ws in wb.worksheets]:
            # We'd need to track data ranges — simplified for now
            pass

    # Remove default empty sheet if we have others
    if len(wb.worksheets) > 1 and default_sheet.title == "Sheet" and default_sheet.max_row <= 1:
        wb.remove(default_sheet)

    # Footer on all sheets
    from openpyxl.styles import Font
    for ws in wb.worksheets:
        footer_row = ws.max_row + 2
        footer_cell = ws.cell(row=footer_row, column=1,
                              value=f"{EXPORT_BRAND_NAME} | {EXPORT_AGENT_NAME} v{APP_VERSION}")
        footer_cell.font = Font(name=FALLBACK_FONT, size=8, color="999999", italic=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
