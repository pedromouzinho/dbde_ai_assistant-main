# =============================================================================
# tests/test_xlsx_engine.py — Tests for Advanced Excel Engine
# =============================================================================

import io
import json
import pytest
from unittest.mock import AsyncMock, patch, MagicMock

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _sample_data(n: int = 10):
    """Generate sample tabular data."""
    return [
        {"id": str(i), "title": f"Item {i}", "state": "Active" if i % 2 == 0 else "Closed",
         "effort": str(i * 5), "created_date": f"2025-01-{i:02d}"}
        for i in range(1, n + 1)
    ]


def _sample_columns():
    return ["id", "title", "state", "effort", "created_date"]


def _simple_spec(n: int = 10):
    return {
        "sheets": [{
            "name": "Dados",
            "title": "Test Data",
            "columns": _sample_columns(),
            "data": _sample_data(n),
            "formulas": True,
            "auto_filter": True,
            "conditional": False,
            "chart": None,
        }],
        "summary": None,
        "charts": [],
    }


def _load_workbook(buf: io.BytesIO):
    from openpyxl import load_workbook
    buf.seek(0)
    return load_workbook(buf)


# =============================================================================
# DATA TYPE DETECTION
# =============================================================================

class TestDataTypeDetection:
    def test_detect_numbers(self):
        from xlsx_engine import detect_column_type
        assert detect_column_type(["10", "20", "30", "40.5"]) == "number"

    def test_detect_dates(self):
        from xlsx_engine import detect_column_type
        assert detect_column_type(["2025-01-01", "2025-02-15", "2025-03-20"]) == "date"

    def test_detect_dates_by_name(self):
        from xlsx_engine import detect_column_type
        assert detect_column_type(["foo", "bar"], col_name="created_date") == "date"

    def test_detect_percentages(self):
        from xlsx_engine import detect_column_type
        assert detect_column_type(["10%", "20.5%", "30%"]) == "percent"

    def test_detect_urls(self):
        from xlsx_engine import detect_column_type
        assert detect_column_type(["https://example.com", "https://test.org"]) == "url"

    def test_detect_url_by_name(self):
        from xlsx_engine import detect_column_type
        assert detect_column_type(["foo"], col_name="url") == "url"

    def test_detect_currency(self):
        from xlsx_engine import detect_column_type
        assert detect_column_type(["€10.50", "€20", "€30.99"]) == "currency"

    def test_detect_text(self):
        from xlsx_engine import detect_column_type
        assert detect_column_type(["hello", "world", "foo"]) == "text"

    def test_empty_values(self):
        from xlsx_engine import detect_column_type
        assert detect_column_type([]) == "text"

    def test_mixed_defaults_to_text(self):
        from xlsx_engine import detect_column_type
        # Mix of types → text
        assert detect_column_type(["10", "hello", "2025-01-01", "30%", "foo"]) == "text"


class TestParseNumber:
    def test_integer(self):
        from xlsx_engine import _parse_number
        assert _parse_number("42") == 42.0

    def test_decimal_dot(self):
        from xlsx_engine import _parse_number
        assert _parse_number("3.14") == 3.14

    def test_decimal_comma(self):
        from xlsx_engine import _parse_number
        assert _parse_number("3,14") == 3.14

    def test_european_format(self):
        from xlsx_engine import _parse_number
        assert _parse_number("1.234,56") == 1234.56

    def test_currency(self):
        from xlsx_engine import _parse_number
        assert _parse_number("€ 100") == 100.0

    def test_percent(self):
        from xlsx_engine import _parse_number
        assert _parse_number("75%") == 75.0

    def test_invalid(self):
        from xlsx_engine import _parse_number
        assert _parse_number("hello") is None

    def test_empty(self):
        from xlsx_engine import _parse_number
        assert _parse_number("") is None


class TestParseDate:
    def test_iso_date(self):
        from xlsx_engine import _parse_date
        dt = _parse_date("2025-01-15")
        assert dt is not None
        assert dt.year == 2025 and dt.month == 1 and dt.day == 15

    def test_iso_datetime(self):
        from xlsx_engine import _parse_date
        dt = _parse_date("2025-01-15T10:30:00")
        assert dt is not None
        assert dt.hour == 10

    def test_european_date(self):
        from xlsx_engine import _parse_date
        dt = _parse_date("15/01/2025")
        assert dt is not None
        assert dt.day == 15

    def test_invalid(self):
        from xlsx_engine import _parse_date
        assert _parse_date("not a date") is None


# =============================================================================
# RENDERER (Layer 3)
# =============================================================================

class TestGenerateWorkbook:
    def test_minimal_workbook(self):
        from xlsx_engine import generate_workbook
        spec = _simple_spec(5)
        buf = generate_workbook(spec)
        wb = _load_workbook(buf)
        assert len(wb.worksheets) >= 1

    def test_data_rows_written(self):
        from xlsx_engine import generate_workbook
        spec = _simple_spec(10)
        buf = generate_workbook(spec)
        wb = _load_workbook(buf)
        ws = wb.worksheets[0]
        # Row 4 = headers, rows 5-14 = data
        assert ws.cell(row=4, column=1).value is not None  # header
        assert ws.cell(row=5, column=1).value is not None  # first data row

    def test_branding_title(self):
        from xlsx_engine import generate_workbook, BRAND_ACCENT_HEX
        spec = _simple_spec(3)
        buf = generate_workbook(spec)
        wb = _load_workbook(buf)
        ws = wb.worksheets[0]
        # Row 1 should contain the title with agent name
        title_val = ws.cell(row=1, column=1).value
        assert title_val is not None
        assert "Test Data" in title_val

    def test_auto_filter_applied(self):
        from xlsx_engine import generate_workbook
        spec = _simple_spec(5)
        buf = generate_workbook(spec)
        wb = _load_workbook(buf)
        ws = wb.worksheets[0]
        assert ws.auto_filter.ref is not None

    def test_freeze_panes(self):
        from xlsx_engine import generate_workbook
        spec = _simple_spec(5)
        buf = generate_workbook(spec)
        wb = _load_workbook(buf)
        ws = wb.worksheets[0]
        assert ws.freeze_panes == "A5"

    def test_formulas_added(self):
        from xlsx_engine import generate_workbook
        spec = _simple_spec(5)
        buf = generate_workbook(spec)
        wb = _load_workbook(buf)
        ws = wb.worksheets[0]
        # Row after data should have formulas
        # Data rows 5-9, formula at row 10
        formula_row = 10
        # Find a cell with a formula (effort column = col 4)
        found_formula = False
        for row in range(10, 15):
            for col in range(1, 6):
                val = ws.cell(row=row, column=col).value
                if isinstance(val, str) and val.startswith("="):
                    found_formula = True
                    break
            if found_formula:
                break
        assert found_formula, "Expected at least one formula row after data"

    def test_multi_sheet(self):
        from xlsx_engine import generate_workbook
        spec = {
            "sheets": [
                {"name": "Sheet 1", "title": "First", "columns": ["a", "b"],
                 "data": [{"a": "1", "b": "2"}], "formulas": False, "auto_filter": True},
                {"name": "Sheet 2", "title": "Second", "columns": ["x", "y"],
                 "data": [{"x": "3", "y": "4"}], "formulas": False, "auto_filter": True},
            ],
            "summary": None,
            "charts": [],
        }
        buf = generate_workbook(spec)
        wb = _load_workbook(buf)
        assert len(wb.worksheets) >= 2

    def test_summary_sheet(self):
        from xlsx_engine import generate_workbook
        spec = _simple_spec(5)
        spec["summary"] = {
            "name": "Resumo",
            "title": "Dashboard",
            "kpis": [
                {"value": "42", "label": "Total", "description": "Items totais"},
                {"value": "85%", "label": "Concluídos"},
            ],
            "sections": [
                {"title": "Destaques", "items": ["Item A", "Item B"]},
            ],
        }
        buf = generate_workbook(spec)
        wb = _load_workbook(buf)
        sheet_names = [ws.title for ws in wb.worksheets]
        assert "Resumo" in sheet_names

    def test_chart_embedded(self):
        from xlsx_engine import generate_workbook
        spec = {
            "sheets": [{
                "name": "Dados",
                "title": "Test",
                "columns": ["category", "value"],
                "data": [
                    {"category": "A", "value": "10"},
                    {"category": "B", "value": "20"},
                    {"category": "C", "value": "30"},
                ],
                "formulas": True,
                "auto_filter": True,
                "chart": {
                    "type": "bar",
                    "title": "Values by Category",
                    "label_column": "category",
                    "value_columns": ["value"],
                    "embed_in_data_sheet": True,
                },
            }],
            "summary": None,
            "charts": [],
        }
        buf = generate_workbook(spec)
        wb = _load_workbook(buf)
        ws = wb.worksheets[0]
        # Check that a chart was added
        assert len(ws._charts) >= 1

    def test_chart_separate_sheet(self):
        from xlsx_engine import generate_workbook
        spec = {
            "sheets": [{
                "name": "Dados",
                "title": "Test",
                "columns": ["category", "value"],
                "data": [
                    {"category": "A", "value": "10"},
                    {"category": "B", "value": "20"},
                ],
                "formulas": True,
                "auto_filter": True,
                "chart": {
                    "type": "pie",
                    "title": "Distribution",
                    "label_column": "category",
                    "value_columns": ["value"],
                    "embed_in_data_sheet": False,
                },
            }],
            "summary": None,
            "charts": [],
        }
        buf = generate_workbook(spec)
        wb = _load_workbook(buf)
        assert len(wb.worksheets) >= 2  # data + chart sheet

    def test_native_number_types(self):
        from xlsx_engine import generate_workbook
        spec = {
            "sheets": [{
                "name": "Numbers",
                "title": "Test Numbers",
                "columns": ["label", "amount"],
                "data": [
                    {"label": "Item A", "amount": "1500"},
                    {"label": "Item B", "amount": "2300.50"},
                ],
                "formulas": True,
                "auto_filter": True,
            }],
            "summary": None,
            "charts": [],
        }
        buf = generate_workbook(spec)
        wb = _load_workbook(buf)
        ws = wb.worksheets[0]
        # Amount column (col 2), row 5 should be numeric
        val = ws.cell(row=5, column=2).value
        assert isinstance(val, (int, float)), f"Expected numeric, got {type(val)}: {val}"

    def test_url_hyperlinks(self):
        from xlsx_engine import generate_workbook
        spec = {
            "sheets": [{
                "name": "Links",
                "title": "URL Test",
                "columns": ["name", "url"],
                "data": [
                    {"name": "Google", "url": "https://google.com"},
                ],
                "formulas": False,
                "auto_filter": True,
            }],
            "summary": None,
            "charts": [],
        }
        buf = generate_workbook(spec)
        wb = _load_workbook(buf)
        ws = wb.worksheets[0]
        url_cell = ws.cell(row=5, column=2)
        assert url_cell.hyperlink is not None

    def test_empty_sheets_removed(self):
        from xlsx_engine import generate_workbook
        spec = {
            "sheets": [
                {"name": "Empty", "title": "Empty", "columns": [], "data": []},
                {"name": "Full", "title": "Full", "columns": ["a"],
                 "data": [{"a": "1"}], "formulas": False, "auto_filter": True},
            ],
            "summary": None,
            "charts": [],
        }
        buf = generate_workbook(spec)
        wb = _load_workbook(buf)
        sheet_names = [ws.title for ws in wb.worksheets]
        assert "Empty" not in sheet_names
        assert "Full" in sheet_names

    def test_conditional_formatting(self):
        from xlsx_engine import generate_workbook
        spec = {
            "sheets": [{
                "name": "CF",
                "title": "Conditional",
                "columns": ["name", "score"],
                "data": [
                    {"name": "A", "score": "10"},
                    {"name": "B", "score": "50"},
                    {"name": "C", "score": "90"},
                ],
                "formulas": True,
                "auto_filter": True,
                "conditional": True,
            }],
            "summary": None,
            "charts": [],
        }
        buf = generate_workbook(spec)
        wb = _load_workbook(buf)
        ws = wb.worksheets[0]
        assert len(ws.conditional_formatting) > 0

    def test_line_chart(self):
        from xlsx_engine import generate_workbook
        spec = {
            "sheets": [{
                "name": "Trend",
                "title": "Trend",
                "columns": ["month", "revenue"],
                "data": [
                    {"month": "Jan", "revenue": "100"},
                    {"month": "Feb", "revenue": "150"},
                    {"month": "Mar", "revenue": "200"},
                ],
                "formulas": True,
                "auto_filter": True,
                "chart": {
                    "type": "line",
                    "title": "Revenue Trend",
                    "label_column": "month",
                    "value_columns": ["revenue"],
                    "embed_in_data_sheet": True,
                },
            }],
            "summary": None,
            "charts": [],
        }
        buf = generate_workbook(spec)
        wb = _load_workbook(buf)
        ws = wb.worksheets[0]
        assert len(ws._charts) >= 1


# =============================================================================
# VALIDATION (Layer 2)
# =============================================================================

class TestValidation:
    def test_removes_empty_sheets(self):
        from xlsx_engine import _validate_and_fix_workbook
        spec = {
            "sheets": [
                {"name": "Empty", "data": [], "columns": ["a"]},
                {"name": "Full", "data": [{"a": "1"}], "columns": ["a"]},
            ],
            "charts": [],
        }
        result = _validate_and_fix_workbook(spec)
        assert len(result["sheets"]) == 1
        assert result["sheets"][0]["name"] == "Full"

    def test_splits_large_sheets(self):
        from xlsx_engine import _validate_and_fix_workbook, MAX_ROWS_PER_SHEET
        data = [{"a": str(i)} for i in range(MAX_ROWS_PER_SHEET + 100)]
        spec = {
            "sheets": [{"name": "Big", "data": data, "columns": ["a"]}],
            "charts": [],
        }
        result = _validate_and_fix_workbook(spec)
        assert len(result["sheets"]) == 2

    def test_trims_columns(self):
        from xlsx_engine import _validate_and_fix_workbook, MAX_COLS
        cols = [f"col_{i}" for i in range(MAX_COLS + 10)]
        spec = {
            "sheets": [{"name": "Wide", "data": [{"col_0": "x"}], "columns": cols}],
            "charts": [],
        }
        result = _validate_and_fix_workbook(spec)
        assert len(result["sheets"][0]["columns"]) == MAX_COLS

    def test_deduplicates_columns(self):
        from xlsx_engine import _validate_and_fix_workbook
        spec = {
            "sheets": [{"name": "Dupes", "data": [{"a": "1"}], "columns": ["a", "a", "b", "b"]}],
            "charts": [],
        }
        result = _validate_and_fix_workbook(spec)
        cols = result["sheets"][0]["columns"]
        assert len(cols) == len(set(cols)), f"Duplicate columns found: {cols}"

    def test_max_sheets(self):
        from xlsx_engine import _validate_and_fix_workbook, MAX_SHEETS
        spec = {
            "sheets": [{"name": f"S{i}", "data": [{"a": "1"}], "columns": ["a"]}
                       for i in range(MAX_SHEETS + 5)],
            "charts": [],
        }
        result = _validate_and_fix_workbook(spec)
        assert len(result["sheets"]) <= MAX_SHEETS

    def test_unique_sheet_names(self):
        from xlsx_engine import _validate_and_fix_workbook
        spec = {
            "sheets": [
                {"name": "Data", "data": [{"a": "1"}], "columns": ["a"]},
                {"name": "Data", "data": [{"a": "2"}], "columns": ["a"]},
            ],
            "charts": [],
        }
        result = _validate_and_fix_workbook(spec)
        names = [s["name"] for s in result["sheets"]]
        assert len(names) == len(set(names)), f"Duplicate sheet names: {names}"


# =============================================================================
# SAFE SHEET TITLE
# =============================================================================

class TestSafeSheetTitle:
    def test_normal(self):
        from xlsx_engine import _safe_sheet_title
        assert _safe_sheet_title("My Sheet") == "My Sheet"

    def test_strips_forbidden(self):
        from xlsx_engine import _safe_sheet_title
        result = _safe_sheet_title("Sheet [1]: test?")
        assert "[" not in result
        assert "]" not in result
        assert ":" not in result
        assert "?" not in result

    def test_max_length(self):
        from xlsx_engine import _safe_sheet_title
        result = _safe_sheet_title("A" * 50)
        assert len(result) <= 31

    def test_dedup(self):
        from xlsx_engine import _safe_sheet_title
        existing = ["Sheet"]
        result = _safe_sheet_title("Sheet", existing)
        assert result != "Sheet"
        assert result not in existing

    def test_empty(self):
        from xlsx_engine import _safe_sheet_title
        assert _safe_sheet_title("") == "Sheet"


# =============================================================================
# FALLBACK
# =============================================================================

class TestFallback:
    def test_fallback_from_data(self):
        from xlsx_engine import _fallback_workbook_from_data
        data = _sample_data(5)
        spec = _fallback_workbook_from_data(data, title="Test")
        assert len(spec["sheets"]) == 1
        assert spec["sheets"][0]["name"] == "Test"
        assert len(spec["sheets"][0]["data"]) == 5

    def test_fallback_empty(self):
        from xlsx_engine import _fallback_workbook_from_data
        spec = _fallback_workbook_from_data([], title="Empty")
        assert len(spec["sheets"]) == 0

    def test_fallback_infers_columns(self):
        from xlsx_engine import _fallback_workbook_from_data
        data = [{"x": "1", "y": "2"}]
        spec = _fallback_workbook_from_data(data)
        assert spec["sheets"][0]["columns"] == ["x", "y"]

    def test_fallback_generates_valid_workbook(self):
        from xlsx_engine import _fallback_workbook_from_data, generate_workbook
        data = _sample_data(5)
        spec = _fallback_workbook_from_data(data, title="Test", columns=_sample_columns())
        buf = generate_workbook(spec)
        wb = _load_workbook(buf)
        assert len(wb.worksheets) >= 1


# =============================================================================
# TOOL INTEGRATION
# =============================================================================

class TestToolIntegration:
    @pytest.mark.asyncio
    async def test_tool_rejects_empty(self):
        from tools_export import tool_generate_spreadsheet
        result = await tool_generate_spreadsheet(title="Test")
        assert "error" in result

    @pytest.mark.asyncio
    async def test_tool_with_sheets(self):
        from tools_export import tool_generate_spreadsheet
        sheets = [{
            "name": "Data",
            "title": "Test",
            "columns": ["a", "b"],
            "data": [{"a": "1", "b": "2"}, {"a": "3", "b": "4"}],
            "formulas": True,
            "auto_filter": True,
        }]
        result = await tool_generate_spreadsheet(
            title="Test Report", sheets=sheets,
        )
        assert result.get("spreadsheet_generated") is True
        assert result.get("format") == "xlsx"
        assert "_file_download" in result
        assert result["_file_download"]["filename"].endswith(".xlsx")
        assert result["planning_model"] == "structured_input"

    @pytest.mark.asyncio
    async def test_tool_with_content_fallback(self):
        """When Opus fails, falls back to parsing content as JSON."""
        from tools_export import tool_generate_spreadsheet
        content = json.dumps([{"name": "A", "score": "10"}, {"name": "B", "score": "20"}])

        with patch("tools_export.logging"):
            result = await tool_generate_spreadsheet(
                title="Fallback Test", content=content,
            )
        # Should work via fallback (JSON parse)
        assert result.get("spreadsheet_generated") is True or "error" in result

    @pytest.mark.asyncio
    async def test_tool_download_metadata(self):
        from tools_export import tool_generate_spreadsheet
        sheets = [{
            "name": "D",
            "columns": ["x"],
            "data": [{"x": "1"}],
        }]
        result = await tool_generate_spreadsheet(title="Meta Test", sheets=sheets)
        assert result.get("spreadsheet_generated") is True
        dl = result["_file_download"]
        assert dl["format"] == "xlsx"
        assert dl["primary"] is True
        assert "download_id" in dl
        assert dl["expires_in_seconds"] > 0
