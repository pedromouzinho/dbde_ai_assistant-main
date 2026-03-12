import base64
import csv
import io
import json
import re

import openpyxl
import pytest

import tools_email


class _DownloadCapture:
    def __init__(self):
        self.entries = []

    async def __call__(self, content, mime_type, filename, fmt, **kwargs):
        self.entries.append(
            {
                "content": bytes(content),
                "mime_type": mime_type,
                "filename": filename,
                "format": fmt,
                "kwargs": dict(kwargs),
            }
        )
        return f"download-{len(self.entries)}"

    def find(self, suffix):
        for entry in self.entries:
            if entry["filename"].endswith(suffix):
                return entry
        raise AssertionError(f"Download com suffix {suffix} não encontrado")


def _workbook_bytes(headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_message_input_text_extracts_known_fields():
    text = (
        "EntryID: AAA123 | Subject: Pedido urgente | From: cliente@empresa.pt | "
        "Body: Preciso de ajuda hoje | Importance: High"
    )
    parsed = tools_email._parse_message_input_text(text)
    assert parsed["entryid"] == "AAA123"
    assert parsed["subject"] == "Pedido urgente"
    assert parsed["from"] == "cliente@empresa.pt"
    assert parsed["body"] == "Preciso de ajuda hoje"
    assert parsed["importance"] == "High"


def test_canonicalize_email_record_from_message_input():
    row = {
        "MessageInput": (
            "EntryID: ID-1 | StoreID: STORE-1 | Subject: Falha no portal | "
            "From: cliente@empresa.pt | Body: o site caiu | ReceivedTime: 2026-03-08T10:00:00Z"
        )
    }
    normalized = tools_email._canonicalize_email_record(row, "1")
    assert normalized["entry_id"] == "ID-1"
    assert normalized["store_id"] == "STORE-1"
    assert normalized["subject"] == "Falha no portal"
    assert normalized["from"] == "cliente@empresa.pt"
    assert normalized["body"] == "o site caiu"


def test_canonicalize_email_record_from_flat_columns():
    row = {
        "EntryID": "ID-2",
        "Subject": "Extrato mensal",
        "SenderName": "Banco XPTO",
        "From": "noreply@banco.pt",
        "Body": "Segue em anexo.",
    }
    normalized = tools_email._canonicalize_email_record(row, "2")
    assert normalized["entry_id"] == "ID-2"
    assert normalized["subject"] == "Extrato mensal"
    assert normalized["sender_name"] == "Banco XPTO"
    assert normalized["from"] == "noreply@banco.pt"


def test_outlook_action_script_contains_expected_operations():
    script = tools_email._build_outlook_actions_powershell("emails.csv")
    assert "GetItemFromID" in script
    assert ".Move(" in script
    assert ".Categories" in script
    assert ".MarkAsTask(" in script
    assert "emails.csv" in script


@pytest.mark.asyncio
async def test_prepare_outlook_draft_generates_msg_launcher(monkeypatch):
    capture = _DownloadCapture()
    monkeypatch.setattr(tools_email, "_store_generated_file", capture)

    result = await tools_email.tool_prepare_outlook_draft(
        subject="Resposta ao cliente",
        body="<p>Olá,<br>Segue resposta.</p>",
        to="cliente@empresa.pt; apoio@empresa.pt",
        cc="gestor@empresa.pt",
    )

    assert result["status"] == "ok"
    assert len(result["_auto_file_downloads"]) == 1
    assert result["_auto_file_downloads"][0]["label"] == "Gerar .msg e abrir draft no Outlook (.cmd)"
    assert result["_auto_file_downloads"][0]["primary"] is True
    launcher = capture.find(".cmd")["content"].decode("utf-8")
    encoded_command = re.search(r"-EncodedCommand ([A-Za-z0-9+/=]+)", launcher)
    assert encoded_command is not None
    script = base64.b64decode(encoded_command.group(1)).decode("utf-16le")
    payload_b64 = re.search(r"FromBase64String\('([^']+)'\)", script)
    assert payload_b64 is not None
    payload = json.loads(base64.b64decode(payload_b64.group(1)).decode("utf-8"))
    assert payload["to"] == "cliente@empresa.pt; apoio@empresa.pt"
    assert "ExecutionPolicy Bypass" in launcher
    assert "DBDE_DRAFT_DIR" in launcher
    assert "CreateItem(0)" in script
    assert ".SaveAs(" in script
    assert "Resposta ao cliente.msg" in script
    assert ".Display()" in script


@pytest.mark.asyncio
async def test_classify_uploaded_emails_from_csv_generates_outlook_pack(monkeypatch):
    capture = _DownloadCapture()
    monkeypatch.setattr(tools_email, "_store_generated_file", capture)

    async def fake_table_query(*args, **kwargs):
        return [
            {
                "Filename": "emails.csv",
                "RawBlobRef": "container/blob.csv",
                "UploadedAt": "2026-03-08T10:00:00+00:00",
                "UserSub": "tester",
            }
        ]

    async def fake_blob_download_bytes(container, blob_name):
        assert container == "container"
        assert blob_name == "blob.csv"
        return (
            "EntryID,Subject,From,Body\n"
            "ID-1,URGENTE: pagamento bloqueado,cliente@empresa.pt,Preciso resolver hoje\n"
            "ID-2,Newsletter,noreply@empresa.pt,Resumo semanal\n"
        ).encode("utf-8")

    async def fake_llm_simple(prompt, tier="standard", max_tokens=0, response_format=None):
        assert response_format is not None
        if "ID-1" in prompt:
            return json.dumps(
                {
                    "decisions": [
                        {
                            "row_id": "1",
                            "label": "Urgente",
                            "confidence": 0.97,
                            "reason": "Assunto e body indicam urgência operacional.",
                            "summary": "Pagamento bloqueado urgente",
                            "requires_manual_review": False,
                        },
                        {
                            "row_id": "2",
                            "label": "FYI",
                            "confidence": 0.91,
                            "reason": "Conteúdo informativo sem pedido de ação.",
                            "summary": "Newsletter",
                            "requires_manual_review": False,
                        },
                    ]
                }
            )
        raise AssertionError("Prompt inesperado")

    monkeypatch.setattr(tools_email, "table_query", fake_table_query)
    monkeypatch.setattr(tools_email, "blob_download_bytes", fake_blob_download_bytes)
    monkeypatch.setattr(tools_email, "llm_simple", fake_llm_simple)

    result = await tools_email.tool_classify_uploaded_emails(
        instructions="Marca como urgente emails com bloqueios operacionais e o resto como FYI.",
        conv_id="conv-1",
        user_sub="tester",
        label_actions=[
            {"label": "Urgente", "action_type": "flag", "target": "today"},
            {"label": "FYI", "action_type": "none", "target": ""},
        ],
    )

    assert result["status"] == "ok"
    assert result["counts_by_label"] == {"Urgente": 1, "FYI": 1}
    assert len(result["_auto_file_downloads"]) == 4
    assert result["_auto_file_downloads"][0]["label"] == "Aplicar ações no Outlook (.ps1)"
    assert result["_auto_file_downloads"][0]["primary"] is True

    xlsx_entry = capture.find(".xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_entry["content"]))
    assert wb.sheetnames == ["Output", "Actions", "Config"]
    headers = [cell.value for cell in wb["Output"][1]]
    assert "Label" in headers
    assert "ActionType" in headers
    actions_headers = [cell.value for cell in wb["Actions"][1]]
    assert actions_headers[:4] == ["RowId", "EntryID", "StoreID", "Subject"]

    csv_entry = capture.find(".csv")
    rows = list(csv.DictReader(io.StringIO(csv_entry["content"].decode("utf-8-sig"))))
    assert rows[0]["Label"] == "Urgente"
    assert rows[0]["ActionType"] == "flag"
    assert rows[1]["Label"] == "FYI"

    manifest = json.loads(capture.find(".json")["content"].decode("utf-8"))
    assert manifest["counts_by_label"] == {"Urgente": 1, "FYI": 1}
    assert manifest["actions"][0]["EntryID"] == "ID-1"


@pytest.mark.asyncio
async def test_classify_uploaded_emails_supports_legacy_messageinput_xlsx(monkeypatch):
    capture = _DownloadCapture()
    monkeypatch.setattr(tools_email, "_store_generated_file", capture)

    workbook_content = _workbook_bytes(
        ["MessageInput"],
        [
            [
                "EntryID: LEG-1 | Subject: Pedido com anexo | From: cliente@empresa.pt | "
                "Body: envio screenshots para rever"
            ]
        ],
    )

    async def fake_table_query(*args, **kwargs):
        return [
            {
                "Filename": "AgenteDeEmails.xlsx",
                "RawBlobRef": "container/legacy.xlsx",
                "UploadedAt": "2026-03-08T10:00:00+00:00",
            }
        ]

    async def fake_blob_download_bytes(container, blob_name):
        return workbook_content

    async def fake_llm_simple(prompt, tier="standard", max_tokens=0, response_format=None):
        return json.dumps(
            {
                "decisions": [
                    {
                        "row_id": "1",
                        "label": "Screenshots",
                        "confidence": 0.88,
                        "reason": "Refere screenshots no corpo.",
                        "summary": "Email com screenshots",
                        "requires_manual_review": True,
                    }
                ]
            }
        )

    monkeypatch.setattr(tools_email, "table_query", fake_table_query)
    monkeypatch.setattr(tools_email, "blob_download_bytes", fake_blob_download_bytes)
    monkeypatch.setattr(tools_email, "llm_simple", fake_llm_simple)

    result = await tools_email.tool_classify_uploaded_emails(
        instructions="Emails com screenshots vão para revisão manual.",
        conv_id="conv-legacy",
        label_actions=[{"label": "Screenshots", "action_type": "move", "target": "Inbox/Screenshots"}],
    )

    assert result["status"] == "ok"
    assert result["counts_by_label"] == {"Screenshots": 1}
    csv_entry = capture.find(".csv")
    rows = list(csv.DictReader(io.StringIO(csv_entry["content"].decode("utf-8-sig"))))
    assert rows[0]["EntryID"] == "LEG-1"
    assert rows[0]["ActionType"] == "move"
    assert rows[0]["ActionTarget"] == "Inbox/Screenshots"


@pytest.mark.asyncio
async def test_classify_uploaded_emails_accepts_tsv_files(monkeypatch):
    capture = _DownloadCapture()
    monkeypatch.setattr(tools_email, "_store_generated_file", capture)

    async def fake_table_query(*args, **kwargs):
        return [{"Filename": "emails.tsv", "RawBlobRef": "container/blob.tsv", "UploadedAt": "2026-03-08T10:00:00+00:00"}]

    async def fake_blob_download_bytes(container, blob_name):
        return "EntryID\tSubject\tBody\nID-1\tTeste TSV\tConteúdo tab\n".encode("utf-8")

    async def fake_llm_simple(prompt, tier="standard", max_tokens=0, response_format=None):
        return json.dumps(
            {
                "decisions": [
                    {
                        "row_id": "1",
                        "label": "review",
                        "confidence": 0.8,
                        "reason": "TSV test.",
                        "summary": "TSV email",
                        "requires_manual_review": False,
                    }
                ]
            }
        )

    monkeypatch.setattr(tools_email, "table_query", fake_table_query)
    monkeypatch.setattr(tools_email, "blob_download_bytes", fake_blob_download_bytes)
    monkeypatch.setattr(tools_email, "llm_simple", fake_llm_simple)

    result = await tools_email.tool_classify_uploaded_emails(
        instructions="Classifica tudo como review.",
        conv_id="conv-tsv",
    )

    assert result["status"] == "ok"
    assert result["counts_by_label"] == {"review": 1}


@pytest.mark.asyncio
async def test_classify_uploaded_emails_uses_default_actions_when_none_provided(monkeypatch):
    capture = _DownloadCapture()
    monkeypatch.setattr(tools_email, "_store_generated_file", capture)

    async def fake_table_query(*args, **kwargs):
        return [{"Filename": "emails.csv", "RawBlobRef": "container/default.csv", "UploadedAt": "2026-03-08T10:00:00+00:00"}]

    async def fake_blob_download_bytes(container, blob_name):
        return "EntryID,Subject,Body\nID-1,FYI,sem urgência\n".encode("utf-8")

    async def fake_llm_simple(prompt, tier="standard", max_tokens=0, response_format=None):
        return json.dumps(
            {
                "decisions": [
                    {
                        "row_id": "1",
                        "label": "review",
                        "confidence": 0.5,
                        "reason": "Sem contexto suficiente.",
                        "summary": "Rever manualmente",
                        "requires_manual_review": True,
                    }
                ]
            }
        )

    monkeypatch.setattr(tools_email, "table_query", fake_table_query)
    monkeypatch.setattr(tools_email, "blob_download_bytes", fake_blob_download_bytes)
    monkeypatch.setattr(tools_email, "llm_simple", fake_llm_simple)

    result = await tools_email.tool_classify_uploaded_emails(
        instructions="Se não houver urgência clara, manda para revisão.",
        conv_id="conv-default",
    )

    assert result["status"] == "ok"
    assert result["label_actions"][0]["label"] == "urgent"
    assert result["label_actions"][1]["label"] == "review"
    assert result["counts_by_label"] == {"review": 1}


@pytest.mark.asyncio
async def test_classify_uploaded_emails_returns_error_without_instructions():
    result = await tools_email.tool_classify_uploaded_emails(instructions="", conv_id="conv-x")
    assert result["error"] == "instructions é obrigatório para classificar emails."


def test_classification_prompt_truncates_long_instructions():
    prompt = tools_email._build_classification_prompt(
        "x" * 5000,
        [{"label": "review", "action_type": "none", "target": ""}],
        "review",
        [{"row_id": "1", "subject": "Teste", "body": "Corpo"}],
    )

    start = prompt.index("<user_instructions>") + len("<user_instructions>\n")
    end = prompt.index("\n</user_instructions>")
    instructions_block = prompt[start:end]

    assert len(instructions_block) == tools_email._EMAIL_CLASSIFICATION_INSTRUCTIONS_LIMIT
